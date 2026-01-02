import os
import argparse
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# Colors for highlighting changes
LIGHT_ORANGE = "FFD699"  # Light orange for updates
LIGHT_GREEN = "90EE90"   # Light green for additions

UPDATE_FILL = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
ADD_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")


def update_cell(sheet, row, col, value):
    """Update a cell value and apply orange highlighting."""
    cell = sheet.cell(row=row, column=col)
    cell.value = value
    cell.fill = UPDATE_FILL


def add_row(sheet, row_num, row_data, is_new_sheet=False):
    """Add a row of data with appropriate highlighting."""
    fill = ADD_FILL if is_new_sheet else ADD_FILL
    for col_idx, value in enumerate(row_data, start=1):
        cell = sheet.cell(row=row_num, column=col_idx)
        cell.value = value
        cell.fill = fill


def process_existing_sheet(workbook, sheet_config):
    """Process updates to an existing sheet."""
    sheet_name = sheet_config.get("name")
    if sheet_name not in workbook.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found in workbook, skipping")
        return False

    sheet = workbook[sheet_name]
    modified = False

    # Process cell updates
    if "updates" in sheet_config:
        for update in sheet_config["updates"]:
            row = update.get("row")
            col = update.get("col")
            value = update.get("value")
            if row and col and value is not None:
                update_cell(sheet, row, col, value)
                modified = True
                print(f"  Updated cell ({row}, {col}) = '{value}'")

    # Process row additions
    if "add_rows" in sheet_config:
        for row_config in sheet_config["add_rows"]:
            start_row = row_config.get("row")
            if not start_row:
                continue

            # Support both single row (data) and multiple rows (rows)
            if "rows" in row_config:
                # Multiple rows starting from the specified row
                for offset, row_data in enumerate(row_config["rows"]):
                    row_num = start_row + offset
                    add_row(sheet, row_num, row_data, is_new_sheet=False)
                    modified = True
                    print(f"  Added row {row_num} with {len(row_data)} cells")
            elif "data" in row_config:
                # Single row (backwards compatible)
                data = row_config.get("data", [])
                if data:
                    add_row(sheet, start_row, data, is_new_sheet=False)
                    modified = True
                    print(f"  Added row {start_row} with {len(data)} cells")

    # Set sheet tab color if modified
    if modified:
        sheet.sheet_properties.tabColor = LIGHT_ORANGE

    return modified


def process_new_sheet(workbook, sheet_config):
    """Create and populate a new sheet."""
    sheet_name = sheet_config.get("name")
    if not sheet_name:
        print("Warning: New sheet missing 'name', skipping")
        return False

    if sheet_name in workbook.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' already exists, skipping new sheet creation")
        return False

    sheet = workbook.create_sheet(title=sheet_name)
    print(f"  Created new sheet '{sheet_name}'")

    # Process row additions
    if "rows" in sheet_config:
        for row_idx, row_data in enumerate(sheet_config["rows"], start=1):
            add_row(sheet, row_idx, row_data, is_new_sheet=True)
            print(f"  Added row {row_idx} with {len(row_data)} cells")

    # Set sheet tab color for new sheet
    sheet.sheet_properties.tabColor = LIGHT_GREEN

    return True


def load_yaml_config(yaml_path):
    """Load the YAML configuration file."""
    with open(yaml_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def generate_output_filename(input_path):
    """Generate output filename by adding '_amended' suffix."""
    directory = os.path.dirname(input_path)
    filename = os.path.basename(input_path)
    name, ext = os.path.splitext(filename)
    output_filename = f"{name}_amended{ext}"
    return os.path.join(directory, output_filename) if directory else output_filename


def main():
    parser = argparse.ArgumentParser(
        prog="update_excel",
        description="Update an Excel workbook according to a YAML configuration file",
        epilog="Updates are highlighted in orange, additions in green",
    )
    parser.add_argument("excel_file", help="The input Excel workbook (.xlsx)")
    parser.add_argument("yaml_file", help="The YAML configuration file")
    parser.add_argument(
        "-o", "--output",
        help="Output filename (default: input filename with '_amended' suffix)",
        default=None
    )
    args = parser.parse_args()

    # Validate input files exist
    if not os.path.exists(args.excel_file):
        print(f"Error: Excel file '{args.excel_file}' not found")
        return 1

    if not os.path.exists(args.yaml_file):
        print(f"Error: YAML file '{args.yaml_file}' not found")
        return 1

    # Load workbook and config
    print(f"\nLoading workbook: {args.excel_file}")
    workbook = load_workbook(args.excel_file)

    print(f"Loading configuration: {args.yaml_file}")
    config = load_yaml_config(args.yaml_file)

    # Process existing sheets
    if "existing_sheets" in config:
        print("\nProcessing existing sheets:")
        for sheet_config in config["existing_sheets"]:
            sheet_name = sheet_config.get("name", "unnamed")
            print(f"\n  Sheet: {sheet_name}")
            process_existing_sheet(workbook, sheet_config)

    # Process new sheets
    if "new_sheets" in config:
        print("\nProcessing new sheets:")
        for sheet_config in config["new_sheets"]:
            process_new_sheet(workbook, sheet_config)

    # Determine output filename
    output_path = args.output if args.output else generate_output_filename(args.excel_file)

    # Save the modified workbook
    print(f"\nSaving to: {output_path}")
    workbook.save(output_path)
    print("Done!")

    return 0


if __name__ == "__main__":
    exit(main())
