import os
import argparse
from openpyxl import load_workbook
from html import escape


def get_sheet_data(sheet):
    """Extract all data from a sheet as a 2D list."""
    data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            value = cell.value
            if value is None:
                value = ""
            row_data.append(value)
        data.append(row_data)

    # Remove trailing empty rows
    while data and all(cell == "" for cell in data[-1]):
        data.pop()

    # Remove trailing empty columns
    if data:
        max_col = 0
        for row in data:
            for i in range(len(row) - 1, -1, -1):
                if row[i] != "":
                    max_col = max(max_col, i + 1)
                    break
        data = [row[:max_col] for row in data]

    return data


def normalize_dimensions(data1, data2):
    """Ensure both datasets have the same dimensions."""
    max_rows = max(len(data1), len(data2))
    max_cols = max(
        max((len(row) for row in data1), default=0),
        max((len(row) for row in data2), default=0)
    )

    def pad_data(data, rows, cols):
        result = []
        for i in range(rows):
            if i < len(data):
                row = list(data[i])
                row.extend([""] * (cols - len(row)))
            else:
                row = [""] * cols
            result.append(row)
        return result

    return pad_data(data1, max_rows, max_cols), pad_data(data2, max_rows, max_cols)


def compare_sheets(data1, data2):
    """Compare two sheets and return diff information."""
    data1, data2 = normalize_dimensions(data1, data2)

    diff_grid = []
    has_differences = False

    for row_idx, (row1, row2) in enumerate(zip(data1, data2)):
        diff_row = []
        row1_empty = all(cell == "" for cell in row1)
        row2_empty = all(cell == "" for cell in row2)

        for col_idx, (cell1, cell2) in enumerate(zip(row1, row2)):
            cell1_str = str(cell1) if cell1 != "" else ""
            cell2_str = str(cell2) if cell2 != "" else ""

            if cell1_str == cell2_str:
                diff_row.append({
                    "type": "unchanged",
                    "value": cell1_str
                })
            elif cell1_str == "" and cell2_str != "":
                # Added in new
                has_differences = True
                diff_row.append({
                    "type": "added",
                    "value": cell2_str
                })
            elif cell1_str != "" and cell2_str == "":
                # Deleted from old
                has_differences = True
                diff_row.append({
                    "type": "deleted",
                    "value": cell1_str
                })
            else:
                # Modified
                has_differences = True
                diff_row.append({
                    "type": "modified",
                    "old_value": cell1_str,
                    "new_value": cell2_str
                })

        # Check if entire row was added or deleted
        row_type = "unchanged"
        if row1_empty and not row2_empty:
            row_type = "added_row"
        elif not row1_empty and row2_empty:
            row_type = "deleted_row"

        diff_grid.append({
            "row_type": row_type,
            "cells": diff_row
        })

    return diff_grid, has_differences


def create_single_sheet_diff(data, diff_type):
    """Create a diff grid for a sheet that exists only in one workbook.

    diff_type: 'added' for new sheets, 'deleted' for removed sheets
    """
    diff_grid = []

    for row_data in data:
        diff_row = []
        for cell in row_data:
            cell_str = str(cell) if cell != "" else ""
            diff_row.append({
                "type": diff_type,
                "value": cell_str
            })

        row_type = "added_row" if diff_type == "added" else "deleted_row"
        diff_grid.append({
            "row_type": row_type,
            "cells": diff_row
        })

    return diff_grid


def generate_html(comparisons, output_file, file1_name, file2_name):
    """Generate the HTML diff report."""

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Diff: {escape(file1_name)} vs {escape(file2_name)}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootswatch@5.3.2/dist/cosmo/bootstrap.min.css" rel="stylesheet">
    <style>
        body {{
            padding: 20px;
        }}
        .diff-table {{
            font-size: 0.85rem;
            border-collapse: collapse;
            width: 100%;
        }}
        .diff-table th, .diff-table td {{
            border: 1px solid #dee2e6;
            padding: 4px 8px;
            vertical-align: top;
            white-space: pre-wrap;
            word-break: break-word;
            max-width: 300px;
        }}
        .diff-table th {{
            background-color: #f8f9fa;
            font-weight: 600;
            text-align: center;
        }}
        .row-header {{
            background-color: #f8f9fa;
            font-weight: 600;
            text-align: center;
            min-width: 50px;
        }}
        /* GitHub-style diff colors */
        .cell-added {{
            background-color: #d4edda;
            color: #155724;
        }}
        .cell-deleted {{
            background-color: #f8d7da;
            color: #721c24;
            text-decoration: line-through;
        }}
        .cell-modified {{
            background-color: #fff3cd;
        }}
        .cell-modified .old-value {{
            background-color: #ffeef0;
            color: #721c24;
            text-decoration: line-through;
            display: block;
            padding: 2px 4px;
            margin-bottom: 2px;
            border-radius: 2px;
        }}
        .cell-modified .new-value {{
            background-color: #e6ffed;
            color: #155724;
            display: block;
            padding: 2px 4px;
            border-radius: 2px;
        }}
        .row-added {{
            background-color: #d4edda;
        }}
        .row-deleted {{
            background-color: #f8d7da;
        }}
        .row-added td {{
            background-color: #d4edda;
            color: #155724;
        }}
        .row-deleted td {{
            background-color: #f8d7da;
            color: #721c24;
        }}
        .table-container {{
            overflow-x: auto;
            margin-bottom: 20px;
        }}
        .nav-tabs {{
            margin-bottom: 20px;
        }}
        .tab-badge {{
            font-size: 0.7rem;
            margin-left: 5px;
        }}
        .no-diff {{
            color: #28a745;
        }}
        .has-diff {{
            color: #dc3545;
        }}
        .sheet-only-in {{
            font-style: italic;
            color: #6c757d;
        }}
        .legend {{
            margin-bottom: 20px;
            padding: 15px;
            background-color: #f8f9fa;
            border-radius: 5px;
        }}
        .legend-item {{
            display: inline-block;
            margin-right: 20px;
            padding: 4px 8px;
            border-radius: 3px;
            font-size: 0.85rem;
        }}
        h1 {{
            margin-bottom: 5px;
        }}
        .subtitle {{
            color: #6c757d;
            margin-bottom: 20px;
        }}
    </style>
</head>
<body>
    <div class="container-fluid">
        <h1>Excel Diff Report</h1>
        <p class="subtitle">
            <strong>Old:</strong> {escape(file1_name)}<br>
            <strong>New:</strong> {escape(file2_name)}
        </p>

        <div class="legend">
            <strong>Legend:</strong>
            <span class="legend-item cell-added">Added</span>
            <span class="legend-item cell-deleted">Deleted</span>
            <span class="legend-item cell-modified">Modified</span>
            <span class="legend-item">Unchanged</span>
        </div>

        <ul class="nav nav-tabs" id="sheetTabs" role="tablist">
'''

    # Generate tabs
    for idx, comp in enumerate(comparisons):
        active = "active" if idx == 0 else ""
        selected = "true" if idx == 0 else "false"
        sheet_name = escape(comp["name"])

        if comp["status"] == "compared":
            badge_class = "has-diff" if comp["has_differences"] else "no-diff"
            badge_text = "changed" if comp["has_differences"] else "identical"
            badge = f'<span class="badge tab-badge {badge_class}">{badge_text}</span>'
        elif comp["status"] == "only_in_old":
            badge = '<span class="badge tab-badge bg-danger">removed</span>'
        else:
            badge = '<span class="badge tab-badge bg-success">added</span>'

        html += f'''            <li class="nav-item" role="presentation">
                <button class="nav-link {active}" id="tab-{idx}" data-bs-toggle="tab"
                        data-bs-target="#content-{idx}" type="button" role="tab"
                        aria-controls="content-{idx}" aria-selected="{selected}">
                    {sheet_name}{badge}
                </button>
            </li>
'''

    html += '''        </ul>

        <div class="tab-content" id="sheetTabContent">
'''

    # Generate tab content
    for idx, comp in enumerate(comparisons):
        active = "show active" if idx == 0 else ""
        sheet_name = escape(comp["name"])

        html += f'''            <div class="tab-pane fade {active}" id="content-{idx}" role="tabpanel" aria-labelledby="tab-{idx}">
'''

        if comp["status"] == "only_in_old":
            html += f'                <p class="sheet-only-in">Sheet "{sheet_name}" was deleted (existed only in old workbook)</p>\n'
        elif comp["status"] == "only_in_new":
            html += f'                <p class="sheet-only-in">Sheet "{sheet_name}" was added (exists only in new workbook)</p>\n'

        if comp["diff_grid"]:
            # Generate diff table
            html += '                <div class="table-container">\n'
            html += '                    <table class="diff-table">\n'

            # Header row with column letters
            if comp["diff_grid"]:
                num_cols = len(comp["diff_grid"][0]["cells"]) if comp["diff_grid"][0]["cells"] else 0
                html += '                        <thead><tr><th></th>'
                for col_idx in range(num_cols):
                    col_letter = get_column_letter(col_idx)
                    html += f'<th>{col_letter}</th>'
                html += '</tr></thead>\n'

            html += '                        <tbody>\n'

            for row_idx, row_data in enumerate(comp["diff_grid"]):
                row_class = ""
                if row_data["row_type"] == "added_row":
                    row_class = "row-added"
                elif row_data["row_type"] == "deleted_row":
                    row_class = "row-deleted"

                html += f'                            <tr class="{row_class}">'
                html += f'<td class="row-header">{row_idx + 1}</td>'

                for cell in row_data["cells"]:
                    if cell["type"] == "unchanged":
                        html += f'<td>{escape(cell["value"])}</td>'
                    elif cell["type"] == "added":
                        html += f'<td class="cell-added">{escape(cell["value"])}</td>'
                    elif cell["type"] == "deleted":
                        html += f'<td class="cell-deleted">{escape(cell["value"])}</td>'
                    elif cell["type"] == "modified":
                        old_val = escape(cell["old_value"])
                        new_val = escape(cell["new_value"])
                        html += f'<td class="cell-modified"><span class="old-value">{old_val}</span><span class="new-value">{new_val}</span></td>'

                html += '</tr>\n'

            html += '                        </tbody>\n'
            html += '                    </table>\n'
            html += '                </div>\n'

        html += '            </div>\n'

    html += '''        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
'''

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)


def get_column_letter(col_idx):
    """Convert column index (0-based) to Excel column letter."""
    result = ""
    col_idx += 1  # Convert to 1-based
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def main():
    parser = argparse.ArgumentParser(
        prog="excel_diff",
        description="Generate an HTML diff report comparing two Excel workbooks",
        epilog="Output is an HTML file viewable in any browser",
    )
    parser.add_argument("file1", help="The first (old) Excel workbook (.xlsx)")
    parser.add_argument("file2", help="The second (new) Excel workbook (.xlsx)")
    parser.add_argument(
        "-o", "--output",
        help="Output HTML filename (default: <new_file>_diff.html)",
        default=None
    )
    args = parser.parse_args()

    # Generate default output filename in the same directory as the source file
    if args.output is None:
        dir_name = os.path.dirname(os.path.abspath(args.file2))
        base_name = os.path.basename(args.file2)
        name, _ = os.path.splitext(base_name)
        args.output = os.path.join(dir_name, f"{name}_diff.html")

    # Validate input files
    if not os.path.exists(args.file1):
        print(f"Error: File '{args.file1}' not found")
        return 1

    if not os.path.exists(args.file2):
        print(f"Error: File '{args.file2}' not found")
        return 1

    print(f"\nLoading workbooks...")
    print(f"  Old: {args.file1}")
    print(f"  New: {args.file2}")

    wb1 = load_workbook(args.file1, data_only=True)
    wb2 = load_workbook(args.file2, data_only=True)

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)

    all_sheets = sheets1 | sheets2
    comparisons = []

    print(f"\nComparing {len(all_sheets)} sheet(s)...")

    for sheet_name in sorted(all_sheets):
        if sheet_name in sheets1 and sheet_name in sheets2:
            # Compare sheets
            data1 = get_sheet_data(wb1[sheet_name])
            data2 = get_sheet_data(wb2[sheet_name])
            diff_grid, has_differences = compare_sheets(data1, data2)

            status = "changed" if has_differences else "identical"
            print(f"  {sheet_name}: {status}")

            comparisons.append({
                "name": sheet_name,
                "status": "compared",
                "has_differences": has_differences,
                "diff_grid": diff_grid
            })
        elif sheet_name in sheets1:
            # Sheet only in old workbook (deleted)
            data = get_sheet_data(wb1[sheet_name])
            diff_grid = create_single_sheet_diff(data, "deleted")
            print(f"  {sheet_name}: only in old (removed)")
            comparisons.append({
                "name": sheet_name,
                "status": "only_in_old",
                "has_differences": True,
                "diff_grid": diff_grid
            })
        else:
            # Sheet only in new workbook (added)
            data = get_sheet_data(wb2[sheet_name])
            diff_grid = create_single_sheet_diff(data, "added")
            print(f"  {sheet_name}: only in new (added)")
            comparisons.append({
                "name": sheet_name,
                "status": "only_in_new",
                "has_differences": True,
                "diff_grid": diff_grid
            })

    # Generate HTML report
    print(f"\nGenerating report: {args.output}")
    generate_html(comparisons, args.output, args.file1, args.file2)
    print("Done!")

    return 0


if __name__ == "__main__":
    exit(main())
